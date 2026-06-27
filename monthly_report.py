import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import glob
import os
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')


def week_to_month(year_week_str):
    year, week = year_week_str.split('-W')
    monday = datetime.strptime(f"{year}-W{int(week):02d}-1", "%Y-W%W-%w")
    return monday.strftime('%Y-%m')


def read_gaoshi_file(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, min_col=2, values_only=True))
    records = []
    for row in rows[1:]:
        week = row[0]
        date = row[1]   # Pvm (datetime)
        store = row[5]  # Toimipaikka
        product = row[10]  # Nimike
        sales_incl = row[11]  # My € sis alv
        comm_rate = row[14]   # Sopimus % (e.g. -18)
        if week and store and product and isinstance(sales_incl, (int, float)):
            records.append({
                'week': week,
                'month': week_to_month(week),
                'date': date,
                'store': store,
                'product': product,
                'sales': sales_incl,
                'comm_rate': abs(comm_rate) / 100.0 if isinstance(comm_rate, (int, float)) else 0.18,
            })
    return records


THIN = Side(border_style='thin', color='8EA9C1')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def hcell(ws, r, c, text, bg='1F4E79', fg='FFFFFF', wrap=True, width=None):
    cell = ws.cell(r, c, text)
    cell.font = Font(bold=True, color=fg)
    cell.fill = PatternFill(fill_type='solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)
    cell.border = BORDER
    return cell

def dcell(ws, r, c, value, bg=None, bold=False, fmt=None, align='left'):
    cell = ws.cell(r, c, value)
    if bg:
        cell.fill = PatternFill(fill_type='solid', fgColor=bg)
    if bold:
        cell.font = Font(bold=True)
    if fmt:
        cell.number_format = fmt
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = BORDER
    return cell


def build_combined_sheet(ws, all_records, months, month_date_ranges, agg_store, all_stores_per_month):
    """
    Combined sheet: detail rows first, then store summary below.
    7 columns throughout: 月份 | 日期范围 | 门店 | 产品 | 销售额 | 佣金 | 净收入
    Summary section reuses same columns, product col shows store name, product col blank.
    """
    ws.title = '门店明细与汇总'
    col_widths = [10, 28, 32, 30, 16, 14, 14]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    DETAIL_FILL = ['EBF3FB', 'FFFFFF']
    SUB_FILL    = 'BDD7EE'
    MONTH_FILL  = 'D9E1F2'
    SEP_FILL    = '2E4057'   # dark separator band

    # ── Section 1: 明细 ─────────────────────────────────────────
    ws.row_dimensions[1].height = 35
    detail_headers = ['月份', '日期范围', '门店 / Store', '产品 / Product',
                      '销售额(€)含税', '佣金(€)', '净收入(€)']
    detail_bg = ['1F4E79', '1F4E79', '1F4E79', '1F4E79', '1F4E79', '7B2C2C', '1F6B3A']
    for ci, (h, bg) in enumerate(zip(detail_headers, detail_bg), 1):
        hcell(ws, 1, ci, h, bg=bg)

    # Aggregate detail: (month, store, product)
    detail = defaultdict(lambda: {'sales': 0.0, 'commission': 0.0, 'dates': set()})
    for rec in all_records:
        key = (rec['month'], rec['store'], rec['product'])
        detail[key]['sales']      += rec['sales']
        detail[key]['commission'] += rec['sales'] * rec['comm_rate']
        if isinstance(rec['date'], datetime):
            detail[key]['dates'].add(rec['date'].date())

    # Group detail keys by (month, store) so we can write a subtotal after each store
    from itertools import groupby
    sorted_keys = sorted(detail.keys())  # sorts by (month, store, product)

    row = 2
    ci_color = 0
    for ms, group_iter in groupby(sorted_keys, key=lambda k: (k[0], k[1])):
        month, store = ms
        group_keys = list(group_iter)
        fill = DETAIL_FILL[ci_color]
        ci_color = 1 - ci_color

        store_s = store_c = 0.0
        store_dr = ''

        for key in group_keys:
            _, _, product = key
            d = detail[key]
            s, c = d['sales'], d['commission']
            dates = sorted(d['dates'])
            dr = f"{dates[0]} ~ {dates[-1]}" if dates else ''
            if not store_dr:
                store_dr = dr

            dcell(ws, row, 1, month,         bg=fill, align='center')
            dcell(ws, row, 2, dr,            bg=fill)
            dcell(ws, row, 3, store,         bg=fill)
            dcell(ws, row, 4, product,       bg=fill)
            dcell(ws, row, 5, round(s,   2), bg=fill, fmt='#,##0.00', align='right')
            dcell(ws, row, 6, round(c,   2), bg=fill, fmt='#,##0.00', align='right')
            dcell(ws, row, 7, round(s-c, 2), bg=fill, fmt='#,##0.00', align='right')
            store_s += s
            store_c += c
            row += 1

        # Store subtotal row in detail section
        dcell(ws, row, 1, month,                       bg=SUB_FILL, bold=True, align='center')
        dcell(ws, row, 2, store_dr,                    bg=SUB_FILL)
        dcell(ws, row, 3, store,                       bg=SUB_FILL, bold=True)
        dcell(ws, row, 4, f'小计（{len(group_keys)} 产品）', bg=SUB_FILL, bold=True, align='center')
        dcell(ws, row, 5, round(store_s,        2),    bg=SUB_FILL, bold=True, fmt='#,##0.00', align='right')
        dcell(ws, row, 6, round(store_c,        2),    bg=SUB_FILL, bold=True, fmt='#,##0.00', align='right')
        dcell(ws, row, 7, round(store_s-store_c,2),    bg=SUB_FILL, bold=True, fmt='#,##0.00', align='right')
        row += 1

    # ── Separator ────────────────────────────────────────────────
    row += 1
    ws.row_dimensions[row].height = 20
    for ci in range(1, 8):
        cell = ws.cell(row, ci, '')
        cell.fill = PatternFill(fill_type='solid', fgColor=SEP_FILL)
    # Label in separator
    ws.cell(row, 1, '▼  门店汇总 Store Summary  ▼').font = Font(bold=True, color='FFFFFF')
    ws.cell(row, 1).fill = PatternFill(fill_type='solid', fgColor=SEP_FILL)
    ws.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center')
    row += 1

    # ── Section 2: 汇总 ─────────────────────────────────────────
    ws.row_dimensions[row].height = 35
    summary_headers = ['月份', '日期范围', '门店 / Store', '',
                       '销售额合计(€)', '佣金合计(€)', '净收入合计(€)']
    summary_bg = ['1F4E79', '1F4E79', '1F4E79', '1F4E79', '1F4E79', '7B2C2C', '1F6B3A']
    for ci, (h, bg) in enumerate(zip(summary_headers, summary_bg), 1):
        hcell(ws, row, ci, h, bg=bg)
    row += 1

    grand_s = grand_c = 0.0
    for month in months:
        stores = sorted(all_stores_per_month[month])
        dr = month_date_ranges.get(month, ('', ''))
        date_range_str = f"{dr[0]} ~ {dr[1]}" if dr[0] else ''
        month_s = month_c = 0.0

        for si, store in enumerate(stores):
            data = agg_store[month].get(store, {'sales': 0.0, 'commission': 0.0})
            s, c = data['sales'], data['commission']
            fill = DETAIL_FILL[si % 2]

            dcell(ws, row, 1, month,           bg=MONTH_FILL, bold=True, align='center')
            dcell(ws, row, 2, date_range_str,  bg=fill)
            dcell(ws, row, 3, store,           bg=fill)
            dcell(ws, row, 4, '',              bg=fill)
            dcell(ws, row, 5, round(s,   2),   bg=fill, fmt='#,##0.00', align='right')
            dcell(ws, row, 6, round(c,   2),   bg=fill, fmt='#,##0.00', align='right')
            dcell(ws, row, 7, round(s-c, 2),   bg=fill, fmt='#,##0.00', align='right')
            month_s += s
            month_c += c
            row += 1

        # Month subtotal
        dcell(ws, row, 1, month,                              bg=SUB_FILL, bold=True, align='center')
        dcell(ws, row, 2, date_range_str,                     bg=SUB_FILL)
        dcell(ws, row, 3, f'小计（{len(stores)} 门店）',      bg=SUB_FILL, bold=True, align='center')
        dcell(ws, row, 4, '',                                  bg=SUB_FILL)
        dcell(ws, row, 5, round(month_s,        2), bg=SUB_FILL, bold=True, fmt='#,##0.00', align='right')
        dcell(ws, row, 6, round(month_c,        2), bg=SUB_FILL, bold=True, fmt='#,##0.00', align='right')
        dcell(ws, row, 7, round(month_s-month_c,2), bg=SUB_FILL, bold=True, fmt='#,##0.00', align='right')
        grand_s += month_s
        grand_c += month_c
        row += 2

    # Grand total
    GRAND_FILL = 'C00000'
    dcell(ws, row, 1, '总计',          bg=GRAND_FILL, bold=True, align='center')
    dcell(ws, row, 2, '',              bg=GRAND_FILL)
    dcell(ws, row, 3, 'GRAND TOTAL',  bg=GRAND_FILL, bold=True, align='center')
    dcell(ws, row, 4, '',              bg=GRAND_FILL)
    for ci, val in zip([5, 6, 7], [grand_s, grand_c, grand_s - grand_c]):
        cell = dcell(ws, row, ci, round(val, 2), bg=GRAND_FILL, bold=True, fmt='#,##0.00', align='right')
        cell.font = Font(bold=True, color='FFFFFF')
    for ci in [1, 2, 3, 4]:
        ws.cell(row, ci).font = Font(bold=True, color='FFFFFF')


def build_product_sheet(ws, months, month_date_ranges, agg_product, all_products_per_month):
    """
    Layout: 月份 | 日期范围 | 产品 | 销售额 | 佣金 | 净收入
    """
    ws.title = '产品月度汇总'
    ws.row_dimensions[1].height = 35

    headers = ['月份', '日期范围', '产品 / Product', '销售额(€)含税', '佣金(€)', '净收入(€)']
    col_widths = [10, 28, 34, 18, 14, 14]
    bg_colors = ['1F4E79', '1F4E79', '1F4E79', '1F4E79', '7B2C2C', '1F6B3A']
    for ci, (h, bg) in enumerate(zip(headers, bg_colors), 1):
        hcell(ws, 1, ci, h, bg=bg)
        ws.column_dimensions[get_column_letter(ci)].width = col_widths[ci-1]

    row = 2
    DATA_FILL = 'FFF2CC'
    ALT_FILL  = 'FFFFFF'
    SUB_FILL  = 'FFE082'
    MONTH_FILL = 'FFF9C4'

    for month in months:
        products = sorted(all_products_per_month[month])
        dr = month_date_ranges.get(month, ('', ''))
        date_range_str = f"{dr[0]} ~ {dr[1]}" if dr[0] else ''

        month_total_s = 0.0
        month_total_c = 0.0

        for pi, prod in enumerate(products):
            data = agg_product[month].get(prod, {'sales': 0.0, 'commission': 0.0})
            s = data['sales']
            c = data['commission']
            n = s - c
            fill = DATA_FILL if pi % 2 == 0 else ALT_FILL

            dcell(ws, row, 1, month, bg=MONTH_FILL, bold=True, align='center')
            dcell(ws, row, 2, date_range_str, bg=fill)
            dcell(ws, row, 3, prod, bg=fill)
            dcell(ws, row, 4, round(s, 2), bg=fill, fmt='#,##0.00', align='right')
            dcell(ws, row, 5, round(c, 2), bg=fill, fmt='#,##0.00', align='right')
            dcell(ws, row, 6, round(n, 2), bg=fill, fmt='#,##0.00', align='right')

            month_total_s += s
            month_total_c += c
            row += 1

        # Subtotal
        dcell(ws, row, 1, month, bg=SUB_FILL, bold=True, align='center')
        dcell(ws, row, 2, date_range_str, bg=SUB_FILL)
        dcell(ws, row, 3, f'小计 Subtotal ({len(products)} 产品)', bg=SUB_FILL, bold=True, align='center')
        dcell(ws, row, 4, round(month_total_s, 2), bg=SUB_FILL, bold=True, fmt='#,##0.00', align='right')
        dcell(ws, row, 5, round(month_total_c, 2), bg=SUB_FILL, bold=True, fmt='#,##0.00', align='right')
        dcell(ws, row, 6, round(month_total_s - month_total_c, 2), bg=SUB_FILL, bold=True, fmt='#,##0.00', align='right')
        row += 2



def generate_monthly_report(input_dir):
    pattern = os.path.join(input_dir, 'Komissiomyynnit*.xlsx')
    files = [f for f in glob.glob(pattern) if '_统计结果' not in os.path.basename(f)]

    if not files:
        print(f"No Gaoshi files found in {input_dir}")
        return

    # Output filename: based on input file(s), append _统计结果
    if len(files) == 1:
        base = os.path.splitext(os.path.basename(files[0]))[0]
        output_file = os.path.join(input_dir, f"{base}_统计结果.xlsx")
    else:
        base = os.path.splitext(os.path.basename(files[0]))[0]
        output_file = os.path.join(input_dir, f"{base}_等{len(files)}份_统计结果.xlsx")

    print(f"Found {len(files)} file(s):")
    all_records = []
    for f in files:
        records = read_gaoshi_file(f)
        all_records.extend(records)
        print(f"  {len(records)} rows from: {os.path.basename(f)}")

    # Aggregations
    months = sorted(set(r['month'] for r in all_records))
    agg_store   = defaultdict(lambda: defaultdict(lambda: {'sales': 0.0, 'commission': 0.0}))
    agg_product = defaultdict(lambda: defaultdict(lambda: {'sales': 0.0, 'commission': 0.0}))
    all_stores_per_month   = defaultdict(set)
    all_products_per_month = defaultdict(set)
    month_dates            = defaultdict(set)  # month -> set of dates

    for rec in all_records:
        m, s, p = rec['month'], rec['store'], rec['product']
        sales = rec['sales']
        comm  = sales * rec['comm_rate']

        agg_store[m][s]['sales']      += sales
        agg_store[m][s]['commission'] += comm
        agg_product[m][p]['sales']      += sales
        agg_product[m][p]['commission'] += comm
        all_stores_per_month[m].add(s)
        all_products_per_month[m].add(p)
        if isinstance(rec['date'], datetime):
            month_dates[m].add(rec['date'].date())

    # Date ranges per month
    month_date_ranges = {}
    for m, dates in month_dates.items():
        sd = min(dates)
        ed = max(dates)
        month_date_ranges[m] = (str(sd), str(ed))

    # Build Excel
    wb = openpyxl.Workbook()
    ws1 = wb.active
    build_combined_sheet(ws1, all_records, months, month_date_ranges, agg_store, all_stores_per_month)

    ws2 = wb.create_sheet()
    build_product_sheet(ws2, months, month_date_ranges, agg_product, all_products_per_month)

    if os.path.exists(output_file):
        try:
            os.remove(output_file)
        except PermissionError:
            print(f"\n[错误] 请先关闭已打开的文件再运行：\n  {output_file}")
            return
    wb.save(output_file)
    print(f"\nReport saved: {output_file}")

    # Console output
    print(f"\n{'='*72}")
    print(f"  月度汇总 / Monthly Summary")
    print(f"{'='*72}")
    grand_s = grand_c = 0.0
    for month in months:
        dr = month_date_ranges.get(month, ('', ''))
        print(f"\n  [{month}]  {dr[0]} ~ {dr[1]}")
        print(f"  {'门店':<38} {'销售额':>12} {'佣金':>10} {'净收入':>12}")
        print(f"  {'-'*74}")
        ms = mt_c = 0.0
        for store in sorted(all_stores_per_month[month]):
            d = agg_store[month][store]
            s, c = d['sales'], d['commission']
            print(f"  {store:<38} €{s:>11.2f} €{c:>9.2f} €{s-c:>11.2f}")
            ms += s; mt_c += c
        print(f"  {'小计':<38} €{ms:>11.2f} €{mt_c:>9.2f} €{ms-mt_c:>11.2f}")
        grand_s += ms; grand_c += mt_c

    print(f"\n  {'总计 GRAND TOTAL':<38} €{grand_s:>11.2f} €{grand_c:>9.2f} €{grand_s-grand_c:>11.2f}")

    print(f"\n{'='*72}")
    print(f"  产品汇总 / Product Summary")
    print(f"{'='*72}")
    for month in months:
        dr = month_date_ranges.get(month, ('', ''))
        print(f"\n  [{month}]  {dr[0]} ~ {dr[1]}")
        print(f"  {'产品':<38} {'销售额':>12} {'佣金':>10} {'净收入':>12}")
        print(f"  {'-'*74}")
        for prod in sorted(all_products_per_month[month]):
            d = agg_product[month][prod]
            s, c = d['sales'], d['commission']
            print(f"  {prod:<38} €{s:>11.2f} €{c:>9.2f} €{s-c:>11.2f}")


if __name__ == '__main__':
    if getattr(sys, 'frozen', False):
        input_dir = os.path.dirname(sys.executable)
    else:
        input_dir = os.path.dirname(os.path.abspath(__file__))
    generate_monthly_report(input_dir)
